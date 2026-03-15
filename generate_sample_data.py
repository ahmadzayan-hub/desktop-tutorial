"""
Script to generate a sample Input_Data.xlsx for the Dubai Tram KPI Dashboard.
Run once: python generate_sample_data.py
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def style_header(cell, bg_color="0098FF", font_color="FFFFFF"):
    cell.font = Font(bold=True, color=font_color, size=11)
    cell.fill = PatternFill("solid", fgColor=bg_color)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def style_cell(cell, bold=False):
    cell.alignment = Alignment(horizontal="center", vertical="center")
    if bold:
        cell.font = Font(bold=True)


def add_border(ws, min_row, max_row, min_col, max_col):
    thin = Side(style="thin")
    for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for cell in row:
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)


def create_input_excel(path="Input_Data.xlsx"):
    wb = openpyxl.Workbook()

    # ─── Sheet 1: Configuration ───────────────────────────────────────────────
    ws_cfg = wb.active
    ws_cfg.title = "Configuration"

    cfg_headers = ["Parameter", "Value", "Notes"]
    for col, h in enumerate(cfg_headers, 1):
        c = ws_cfg.cell(row=1, column=col, value=h)
        style_header(c)

    cfg_data = [
        ("Month/Year",          "Feb-26",           "Reporting period"),
        ("Annual Fee (AED)",    97329403.00,         "Contract annual fee"),
        ("TR Required",         6,                   "Trams required per day"),
        ("Target PMPC (%)",     95,                  "Min PMPC completion %"),
        ("Target TSA (%)",      99,                  "Min TSA availability %"),
        ("Weight PMPC",         0.40,                "PMPC deduction weight"),
        ("Weight TSA",          0.15,                "TSA deduction weight"),
        ("MC Percentage",       0.10,                "Maintenance Component %"),
        ("Limit TSR",           2.83,                "Max TSR incidents/month"),
        ("Limit PDR",           3.50,                "Max PDR failures/month"),
        ("Limit TVMR",          65.17,               "Max TVMR failures/month"),
        ("Limit LER",           4.16,                "Max LER failures/month"),
        ("Deduction Cap PMPC",  0.85,                "PMPC deduction floor"),
        ("Deduction Cap TSA",   0.94,                "TSA deduction floor"),
        ("Penalty TSR 5-10",    2000,                "AED flat fee per incident"),
        ("Penalty TSR 10-30",   20000,               "AED flat fee per incident"),
        ("Penalty TSR 30+",     100000,              "AED flat fee per incident"),
        ("Penalty PDR/TVMR/LER",1000,               "AED flat fee per excess failure"),
    ]
    for r, (param, val, note) in enumerate(cfg_data, 2):
        ws_cfg.cell(row=r, column=1, value=param)
        ws_cfg.cell(row=r, column=2, value=val)
        ws_cfg.cell(row=r, column=3, value=note)

    ws_cfg.column_dimensions["A"].width = 26
    ws_cfg.column_dimensions["B"].width = 18
    ws_cfg.column_dimensions["C"].width = 34
    add_border(ws_cfg, 1, len(cfg_data) + 1, 1, 3)

    # ─── Sheet 2: PMPC Data ───────────────────────────────────────────────────
    ws_pmpc = wb.create_sheet("PMPC Data")

    pmpc_headers = ["Month", "Planned Tasks", "Completed Tasks", "Status", "KPI", "Remarks"]
    for col, h in enumerate(pmpc_headers, 1):
        c = ws_pmpc.cell(row=1, column=col, value=h)
        style_header(c)

    pmpc_rows = [
        ("Feb-26", 1555, 1555, "COMP", "OK", "All PM tasks completed"),
    ]
    for r, row in enumerate(pmpc_rows, 2):
        for col, val in enumerate(row, 1):
            ws_pmpc.cell(row=r, column=col, value=val)

    for col in range(1, 7):
        ws_pmpc.column_dimensions[get_column_letter(col)].width = 20
    add_border(ws_pmpc, 1, len(pmpc_rows) + 1, 1, 6)

    # ─── Sheet 3: TSA Data ────────────────────────────────────────────────────
    ws_tsa = wb.create_sheet("TSA Data")

    tsa_headers = ["Month", "Trams Available", "Trams Required", "TSA (%)", "Status"]
    for col, h in enumerate(tsa_headers, 1):
        c = ws_tsa.cell(row=1, column=col, value=h)
        style_header(c)

    tsa_rows = [("Feb-26", 6, 6, 100.0, "PASS")]
    for r, row in enumerate(tsa_rows, 2):
        for col, val in enumerate(row, 1):
            ws_tsa.cell(row=r, column=col, value=val)

    for col in range(1, 6):
        ws_tsa.column_dimensions[get_column_letter(col)].width = 20
    add_border(ws_tsa, 1, len(tsa_rows) + 1, 1, 5)

    # ─── Sheet 4: Incidents ───────────────────────────────────────────────────
    ws_inc = wb.create_sheet("Incidents")

    inc_headers = [
        "Date", "KPI Type", "Asset ID", "Failure Code",
        "Downtime (Mins)", "Is Excused", "Attributed To", "Description"
    ]
    for col, h in enumerate(inc_headers, 1):
        c = ws_inc.cell(row=1, column=col, value=h)
        style_header(c)

    incidents = [
        # TSR incidents
        ("2026-02-18", "TSR", "T11", "IOS075",  9.07, False, "Maintainer", "T11 Tachometer Failure"),
        # PDR incidents
        ("2026-02-05", "PDR", "PSD-01", "OPN_FAIL", 0,    False, "Maintainer", "Door Stuck – Station 1"),
        ("2026-02-12", "PDR", "PSD-02", "SYS_FAIL",  0,   False, "Maintainer", "Door Controller Fault"),
        # TVMR incidents
        ("2026-02-01", "TVMR", "TVM-01", "BLANK_SCR",   0, False, "Maintainer", "Blank Screen"),
        ("2026-02-02", "TVMR", "TVM-02", "CARD_RDR",    0, False, "Maintainer", "Card Reader Fault"),
        ("2026-02-03", "TVMR", "TVM-03", "PAPER_JAM",   0, False, "Maintainer", "Receipt Paper Jam"),
        ("2026-02-04", "TVMR", "TVM-04", "COIN_MECH",   0, False, "Maintainer", "Coin Mechanism Error"),
        ("2026-02-05", "TVMR", "TVM-05", "BLANK_SCR",   0, False, "Maintainer", "Blank Screen"),
        ("2026-02-06", "TVMR", "TVM-06", "CARD_RDR",    0, False, "Maintainer", "Card Reader Fault"),
        ("2026-02-07", "TVMR", "TVM-07", "NET_ERR",     0, False, "Maintainer", "Network Connectivity Error"),
        ("2026-02-08", "TVMR", "TVM-08", "COIN_MECH",   0, False, "Maintainer", "Coin Jam"),
        ("2026-02-09", "TVMR", "TVM-09", "BLANK_SCR",   0, False, "Maintainer", "Display Failure"),
        ("2026-02-10", "TVMR", "TVM-10", "PAPER_JAM",   0, False, "Maintainer", "Paper Jam"),
        ("2026-02-11", "TVMR", "TVM-11", "CARD_RDR",    0, False, "Maintainer", "Card Reader Error"),
        ("2026-02-12", "TVMR", "TVM-12", "BLANK_SCR",   0, False, "Maintainer", "Screen Freeze"),
        ("2026-02-13", "TVMR", "TVM-13", "NET_ERR",     0, False, "Maintainer", "Network Error"),
        ("2026-02-14", "TVMR", "TVM-14", "COIN_MECH",   0, False, "Maintainer", "Coin Mechanism"),
        ("2026-02-15", "TVMR", "TVM-15", "CARD_RDR",    0, False, "Maintainer", "Card Error"),
        ("2026-02-16", "TVMR", "TVM-16", "BLANK_SCR",   0, False, "Maintainer", "Display Off"),
        ("2026-02-17", "TVMR", "TVM-17", "PAPER_JAM",   0, False, "Maintainer", "Paper Jam"),
        ("2026-02-18", "TVMR", "TVM-18", "COIN_MECH",   0, False, "Maintainer", "Coin Error"),
        ("2026-02-19", "TVMR", "TVM-19", "NET_ERR",     0, False, "Maintainer", "Connectivity Loss"),
        ("2026-02-20", "TVMR", "TVM-20", "CARD_RDR",    0, False, "Maintainer", "Card Reader"),
        ("2026-02-21", "TVMR", "TVM-21", "BLANK_SCR",   0, False, "Maintainer", "Screen Freeze"),
        ("2026-02-22", "TVMR", "TVM-22", "PAPER_JAM",   0, False, "Maintainer", "Paper Jam"),
        # LER incidents
        ("2026-02-05", "LER",  "LIFT-01", "DR_SENSOR",  0, False, "Maintainer", "Door Sensor Fault"),
        ("2026-02-10", "LER",  "ESC-01",  "MTR_FAULT", 120, False, "Maintainer", "Motor Trip"),
        ("2026-02-20", "LER",  "LIFT-02", "CTRL_FAULT",  0, False, "Maintainer", "Control Panel Fault"),
    ]

    for r, row in enumerate(incidents, 2):
        for col, val in enumerate(row, 1):
            ws_inc.cell(row=r, column=col, value=val)

    col_widths = [14, 10, 12, 14, 16, 12, 16, 36]
    for col, w in enumerate(col_widths, 1):
        ws_inc.column_dimensions[get_column_letter(col)].width = w
    add_border(ws_inc, 1, len(incidents) + 1, 1, 8)

    wb.save(path)
    print(f"Created: {path}")


if __name__ == "__main__":
    create_input_excel()
